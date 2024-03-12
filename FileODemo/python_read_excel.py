from openpyxl import Workbook, load_workbook

wb = load_workbook(filename="demoexcel.xlsx")
# sh = wb.active #print from 1st sheet
sh = wb['Demosheet'] #set sheet

print(sh['A3'].value)
print(wb['Demosheet']['A3'].value)
print(sh.cell(2, 2).value)

row_ct = sh.max_row
col_ct = sh.max_column

print(row_ct)
print(col_ct)

for i in range(1, row_ct+1):
    for j in range(1, col_ct+1):
        print(sh.cell(row=i, column=j).value, end = ' ')
    print("\n")