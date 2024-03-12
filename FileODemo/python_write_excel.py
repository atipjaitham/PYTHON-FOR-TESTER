# pip install openpyxl 
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = "Atip Jaitham"

testdata = [["NAME","CITY"],["ATIP","BANGKOK"], ["AITIAN","TAIPEI"], ["LINMAO","BEIJING"], ["EUNWON","SEOUL"]]
for data in testdata:
    ws.append((data))

for i in range(1, 6):
    for j in range(1,5):
        ws.cell(row=i, column=j).value = i+j
'''
this is output
2	3	4	5
3	4	5	6
4	5	6	7
5	6	7	8
6	7	8	9
'''

wb.save("demoexcel.xlsx")

